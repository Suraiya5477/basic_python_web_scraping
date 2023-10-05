from bs4 import BeautifulSoup
import requests
import openpyxl

url = "https://www.linkedin.com/jobs/search?keywords=Software%20Engineer&location=Bangladesh&geoId=106215326&trk=public_jobs_jobs-search-bar_search-submit&position=2&pageNum=0&currentJobId=3598856132"

result = requests.get(url)
doc = BeautifulSoup(result.text, "html.parser")

job_titles = doc.find_all("a", class_="hidden-nested-link")
job_statuses = doc.find_all("span", class_="result-benefits__text")
job_cards = doc.find_all("h3", class_="base-search-card__title")
if job_titles and job_statuses and job_cards:
    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the job titles and statuses to the worksheet
    for i, (job_title, job_status,job_card) in enumerate(zip(job_titles, job_statuses,job_cards)):
        job_title_text = job_title.text.strip()
        job_status_text = job_status.text.strip()
        job_card_text = job_card.text.strip()
        worksheet.cell(row=i+1, column=1, value=job_title_text)
        worksheet.cell(row=i+1, column=2, value=job_status_text)
        worksheet.cell(row=i+1, column=3, value=job_card_text)

    # Save the workbook
    workbook.save("job_titles_and_statuses_and_card.xlsx")
    print("Job titles and statuses saved to job_titles_and_statuses_and_card.xlsx")
else:
    print("Job titles or statuses not found.")